# -*- coding: utf-8 -*-

import openpyxl as x
import sys
import time
import urllib2 as u
from openpyxl.styles import PatternFill
import mechanize as m
from bs4 import BeautifulSoup as BS



#---------------------------VARIABLES-----------------------------------#




#---------------------------FUNCTIONS-----------------------------------#

#Function to handle opening a url and returning a beautifulsoup object with that page's html
def connectToLink(url):
    
    for attempt in range(10):
        try:
            request = m.Request(url)
            request.add_header(hdr, "http://wwwsearch.sourceforge.net/mechanize/")
            response = m.urlopen(url)
            data = response.read()
            _soup = BS(data, 'html.parser')
    
            return _soup
            
            break
        except:
            #print e
            print "Connection to website failed, attempt %s" % attempt + 1
            time.sleep(0.2)
        else:
            break
    else:
        print "Connection couldn't be made. Try again later"
        return 0
    
    
        
#A function that loops through a given <tr> to find the name and email.

def outputNameAndEmail(trTag):
    name = ""
    email = "" 
    
    #The name is in bold in the first <td> tag. 
    #This is a really ugly way to do it but the site's very consistent across teams.
    nameTD = trTag.td
    name = nameTD.b.contents[0]
    emailTD = trTag.td.next_sibling.find_all("tr")
    for tr in emailTD:
        if tr.td.has_attr('colspan'):
            JS = tr.td.a['href']
            JSParts = JS.split("'")
            email = JSParts[3] + "@" + JSParts[1]
    
    return (name, email)
    

#Function to handle the page with all the trainer information
def handleTrainerPage(BSTrainerPage, clubName, WSIndex):

    #This variable contains all the lists of ["name", "email", "function"] format
    listOfPageDataSet = []
    listOfNames = []
    
    
    #Print the current club to the console
    #I overwrite the line with a load of spaces to "clear" it, because there is no other easy way to do that on windows.
    sys.stdout.write("\r                                                               ")
    sys.stdout.write("\rCurrently processing club: " + clubName)
    sys.stdout.flush()
    
    #Added this to prevent "NoneType doesn't have attribute 'table'" errors
    try:
        table = BSTrainerPage.find("div", class_="nisObject nisVereinFunktionaer").table
    except(AttributeError):
        return 0
    
    #Make a list of all tr tags in that table
    trList = table.find_all("tr")

    #Iterate over that list by using an index, so we can manipulate things at a higher index
    #We're looking for the TD tags with the function description of the people in this table
    #So we can see if they're a trainer for a junior team or not
    
    for tr in range(0, len(trList)):
        #Create a temp array to store the output of outputNameAndEmail() in
        outputData = [u'name', u'email', u'function']
    

        #_item is the first td tag under a tr tag 
        _item = trList[tr].td
    
        #If that td tag has the attribute class and the class name is tt:
        if _item.has_attr('class') and _item['class'][0] == u'tt':
        
            #If the trainer is a trainer for junior teams, find their name and email and place them in a list
            if "junior" in _item.string.lower():
                outputData[0], outputData[1] = outputNameAndEmail(trList[tr + 1])
                outputData[2] = _item.string
                
                #Check if the names collected aren't duplicates
                if outputData[0] not in listOfNames and outputData[0] != '-' and outputData[1] != '':
                    #Add them to the list of seen names if they arent and append the whole data block to the output list
                    listOfNames.append(outputData[0])
                    listOfPageDataSet.append(outputData)
    
    if listOfPageDataSet != []:
        return listOfPageDataSet
    else:
        return 0

#Main looping function. Called for every URL using a for loop

def mainSearchAndExtract(url, WSIndex):
    soup = connectToLink(url)
    
    if soup == 0:
        return
    #The teams are spread over three columns that are in the form of <td class="nisTeamListe">
    #Create a list of those to iterate through
    teamColumns = soup.find_all("td", class_="nisTeamListe")
    
    for column in teamColumns:
        for link in column.children:
        
            #Clean up name
            name = link.string.split("(")[0].rstrip()
            link = link['href'] + 'a-tr'
            BSTrainerPage = connectToLink(link)
            _output = handleTrainerPage(BSTrainerPage, name, WSIndex)
            
            #Add name to excel file
            ws[('F' + str(WSIndex))] = name
            
            #Made handleTrainerPage return 0 when there was nothing (containing junior) on the page
            #I check for that here to be able to mark those cells red
            if _output != 0 and _output != '':
                for list in _output:
                    #Add data to correct fields and increment index by one
                    ws[('H' + str(WSIndex))], ws[('I' + str(WSIndex))], ws[('J' + str(WSIndex))] = list
                    WSIndex += 1
                    
            
            else:
                #If the output of the handleTrainerPage function is 0 or an empty list,
                #Fill the "name" cell with red
                ws[('F' + str(WSIndex))].fill = redFill
                WSIndex += 1
            WSIndex += 1
            break   #uncomment to debug, stops after first club of every column
    
    
    return WSIndex
            
            
#---------------------------------MAIN--------------------------------------#
redFill = PatternFill(start_color='FFEE1111',
                   end_color='FFEE1111',
                   fill_type='solid')
WSIndex = 2
fileIndex = 1

hdr = "User-Agent"

with open("source.txt", 'r') as f:
    for line in f:
        
        start_time = time.clock()
        
        #Open a new workbook for every line in the source file
        wb = x.load_workbook("Bestand" + str(fileIndex) + ".xlsx")
        ws = wb.get_sheet_by_name("clubs_te_verrijken")
        
        WSIndex = mainSearchAndExtract(line, WSIndex)
        
        wb.save("Bestand" + str(fileIndex) + ".xlsx")
        fileIndex += 1
        
        name = line.split('/')
        sys.stdout.write("\r                                                               ")
        if "football" in name[2]:
            sys.stdout.write("\rDone Processing " + name[4])
        else:
            sys.stdout.write("\rDone Processing " + name[3])
        sys.stdout.flush()
        
        print("--- %s seconds ---" % (time.clock() - start_time))
    


#Open a new workbook for every link in the source file?







        
    
